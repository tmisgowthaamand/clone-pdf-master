"""
Enhance Excel Formatting Before PDF Conversion
This ensures the Excel file has proper formatting that LibreOffice will preserve
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter


def enhance_excel_formatting(excel_path):
    """
    Enhance Excel file formatting to ensure it converts well to PDF
    - Sets proper column widths
    - Applies borders to all cells
    - Centers headers
    - Ensures proper alignment
    """
    print(f"Enhancing Excel formatting for PDF conversion...")
    
    try:
        # Load workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Define styles
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        header_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        
        # Process all cells
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                # Apply border to all cells
                cell.border = border
                
                # First row is header
                if row_idx == 1:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    # Optional: Add background color to headers
                    # cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                else:
                    cell.font = normal_font
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Auto-detect number columns and right-align them
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif cell.value and str(cell.value).replace('.', '').replace(',', '').replace('-', '').isdigit():
                        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Calculate length
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Set column width (add padding)
            adjusted_width = min(max_length + 3, 50)  # Max 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights for better readability
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20  # Standard row height
        
        # Save enhanced workbook
        wb.save(excel_path)
        print(f"✓ Excel formatting enhanced")
        
        return True
        
    except Exception as e:
        print(f"Warning: Could not enhance Excel formatting: {str(e)}")
        return False


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python enhance_excel_formatting.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        sys.exit(1)
    
    if enhance_excel_formatting(excel_file):
        print(f"✓ SUCCESS: {excel_file}")
    else:
        print(f"✗ FAILED")
        sys.exit(1)
