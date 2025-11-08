"""
Memory-optimized routes for Render.com (512MB limit)
Replace the pdf-to-excel route in app.py with this version
"""

import gc
import os
import tempfile
import shutil
from pathlib import Path
from flask import request, send_file, jsonify
from werkzeug.utils import secure_filename

# Get max pages from environment or default to 3
MAX_PAGES = int(os.environ.get('MAX_PAGES_TO_PROCESS', '3'))

def pdf_to_excel_optimized():
    """Memory-optimized PDF to Excel conversion"""
    print("\n" + "="*60)
    print("PDF TO EXCEL CONVERSION (MEMORY OPTIMIZED)")
    print(f"Max pages to process: {MAX_PAGES}")
    print("="*60)
    
    tmpdir = None
    doc = None
    wb = None
    
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        print(f"✓ File received: {file.filename}")
        
        # Create temp directory
        tmpdir = tempfile.mkdtemp()
        
        # Save uploaded PDF
        filename = secure_filename(file.filename)
        pdf_path = os.path.join(tmpdir, filename)
        file.save(pdf_path)
        
        file_size = os.path.getsize(pdf_path) / 1024
        print(f"File size: {file_size:.2f} KB")
        
        # Import only when needed to save memory
        import fitz
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        # Open PDF
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        pages_to_process = min(MAX_PAGES, total_pages)
        
        print(f"Total pages: {total_pages}")
        print(f"Processing: {pages_to_process} pages (memory limit)")
        
        # Create Excel
        excel_name = Path(filename).stem + '.xlsx'
        excel_path = os.path.join(tmpdir, excel_name)
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Extracted Data'
        
        current_row = 1
        
        # Add header
        ws.cell(row=current_row, column=1, value=f"Extracted from: {filename}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 2
        
        ws.cell(row=current_row, column=1, value=f"Pages processed: {pages_to_process} of {total_pages}")
        ws.cell(row=current_row, column=1).font = Font(italic=True, size=10)
        current_row += 2
        
        # Extract text from pages
        for page_num in range(pages_to_process):
            page = doc[page_num]
            text = page.get_text()
            
            if text.strip():
                # Page header
                ws.cell(row=current_row, column=1, value=f"=== Page {page_num + 1} ===")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                # Add text lines
                lines = text.strip().split('\n')
                for line in lines[:100]:  # Limit lines per page
                    if line.strip():
                        ws.cell(row=current_row, column=1, value=line.strip())
                        current_row += 1
                
                current_row += 1  # Spacing
            
            # Force cleanup after each page
            page = None
            gc.collect()
        
        # Close PDF to free memory
        doc.close()
        doc = None
        gc.collect()
        
        # Save Excel
        wb.save(excel_path)
        wb.close()
        wb = None
        gc.collect()
        
        print(f"✓ Excel created: {excel_name}")
        print(f"✓ Conversion complete")
        
        # Send file
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=excel_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
        
    finally:
        # Cleanup
        if doc:
            try:
                doc.close()
            except:
                pass
        
        if wb:
            try:
                wb.close()
            except:
                pass
        
        if tmpdir and os.path.exists(tmpdir):
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except:
                pass
        
        # Force garbage collection
        gc.collect()
        gc.collect()
        print("Memory cleanup completed")


# To use this in app.py, replace the @app.route('/api/convert/pdf-to-excel', methods=['POST'])
# function with:
#
# @app.route('/api/convert/pdf-to-excel', methods=['POST'])
# def pdf_to_excel():
#     from memory_optimized_routes import pdf_to_excel_optimized
#     return pdf_to_excel_optimized()
