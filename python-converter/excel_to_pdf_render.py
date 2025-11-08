"""
Optimized Excel to PDF Converter for Render.com
Produces professional PDFs like the bank statement example
Works on Linux with LibreOffice
"""

import os
import sys
import subprocess
import tempfile
import shutil
from pathlib import Path

def convert_excel_to_pdf_render(excel_path, output_dir):
    """
    Convert Excel to professional PDF using LibreOffice
    Optimized for Render.com deployment
    Produces high-quality output with proper formatting
    """
    print(f"\n{'='*60}")
    print(f"EXCEL TO PDF - Professional Quality")
    print(f"Input: {os.path.basename(excel_path)}")
    print(f"{'='*60}\n")
    
    try:
        # Find LibreOffice
        if sys.platform == 'win32':
            soffice_paths = [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]
            soffice_exe = None
            for path in soffice_paths:
                if os.path.exists(path):
                    soffice_exe = path
                    break
            
            if not soffice_exe:
                print("ERROR: LibreOffice not found")
                return None
                
            # Kill existing processes on Windows
            try:
                subprocess.run(['taskkill', '/F', '/IM', 'soffice.exe', '/T'], 
                              capture_output=True, timeout=5)
                subprocess.run(['taskkill', '/F', '/IM', 'soffice.bin', '/T'], 
                              capture_output=True, timeout=5)
                import time
                time.sleep(1)
            except:
                pass
        else:
            # Linux/Render.com
            soffice_exe = shutil.which('soffice') or 'soffice'
            if not shutil.which(soffice_exe):
                print("ERROR: LibreOffice not found. Install with: apt-get install libreoffice")
                return None
        
        print(f"Using LibreOffice: {soffice_exe}")
        
        # Output PDF path
        base_name = os.path.splitext(os.path.basename(excel_path))[0]
        pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
        
        # LibreOffice command with optimal settings for professional output
        # These settings produce bank-statement quality PDFs
        cmd = [
            soffice_exe,
            '--headless',                    # No GUI
            '--invisible',                   # Hidden mode
            '--nodefault',                   # No default document
            '--nofirststartwizard',          # Skip wizard
            '--nolockcheck',                 # No lock check
            '--nologo',                      # No splash screen
            '--norestore',                   # Don't restore windows
            '--convert-to', 'pdf:calc_pdf_Export',  # Use Calc PDF export filter
            '--outdir', output_dir,          # Output directory
            excel_path                       # Input file
        ]
        
        print(f"Converting to PDF...")
        print(f"Command: {' '.join(cmd[:8])}...")
        
        # Run conversion
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=120,  # 2 minutes timeout
            encoding='utf-8',
            errors='replace',
            env={
                **os.environ,
                'HOME': tempfile.gettempdir(),  # Set HOME for LibreOffice
            }
        )
        
        if result.returncode != 0:
            print(f"ERROR: LibreOffice conversion failed")
            print(f"Return code: {result.returncode}")
            print(f"STDERR: {result.stderr}")
            print(f"STDOUT: {result.stdout}")
            return None
        
        # Check if PDF was created
        if os.path.exists(pdf_path):
            file_size = os.path.getsize(pdf_path) / 1024
            print(f"✓ PDF created successfully: {os.path.basename(pdf_path)}")
            print(f"✓ File size: {file_size:.2f} KB")
            print(f"{'='*60}\n")
            return pdf_path
        else:
            print(f"ERROR: PDF file not created at: {pdf_path}")
            return None
            
    except subprocess.TimeoutExpired:
        print("ERROR: Conversion timeout (120 seconds)")
        return None
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def convert_csv_to_excel_render(csv_path, output_dir):
    """
    Convert CSV to Excel using pandas
    Preserves formatting and data types
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        print(f"Converting CSV to Excel...")
        
        # Read CSV
        df = pd.read_csv(csv_path, encoding='utf-8', on_bad_lines='skip')
        
        # Create Excel file
        base_name = os.path.splitext(os.path.basename(csv_path))[0]
        excel_path = os.path.join(output_dir, f"{base_name}.xlsx")
        
        # Write to Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Style header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Apply borders to all cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✓ CSV converted to Excel: {os.path.basename(excel_path)}")
        return excel_path
        
    except Exception as e:
        print(f"ERROR converting CSV: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# Test function
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_pdf_render.py <excel_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}")
        sys.exit(1)
    
    output_dir = os.path.dirname(input_file) or '.'
    
    if input_file.lower().endswith('.csv'):
        excel_file = convert_csv_to_excel_render(input_file, output_dir)
        if excel_file:
            pdf_file = convert_excel_to_pdf_render(excel_file, output_dir)
    else:
        pdf_file = convert_excel_to_pdf_render(input_file, output_dir)
    
    if pdf_file:
        print(f"\n✓ SUCCESS: {pdf_file}")
    else:
        print(f"\n✗ FAILED")
        sys.exit(1)
