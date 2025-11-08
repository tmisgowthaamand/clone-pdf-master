"""
FAST & ACCURATE PDF to Excel Converter
Optimized for Render.com (512MB RAM limit)
Uses pdfplumber - faster and more accurate than Camelot
"""

import os
import gc
import tempfile
import shutil
from pathlib import Path
from flask import request, send_file, jsonify
from werkzeug.utils import secure_filename

# Get configuration from environment
MAX_PAGES = int(os.environ.get('MAX_PAGES_TO_PROCESS', '10'))
MAX_FILE_SIZE_MB = int(os.environ.get('MAX_FILE_SIZE_MB', '50'))

def pdf_to_excel_fast():
    """Fast and accurate PDF to Excel conversion using pdfplumber"""
    print("\n" + "="*60)
    print("PDF TO EXCEL - FAST MODE")
    print("="*60)
    
    tmpdir = None
    pdf = None
    
    try:
        # Validate request
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        filename = secure_filename(file.filename)
        if not filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Only PDF files allowed'}), 400
        
        print(f"✓ File: {filename}")
        
        # Create temp directory
        tmpdir = tempfile.mkdtemp()
        pdf_path = os.path.join(tmpdir, filename)
        file.save(pdf_path)
        
        # Check file size
        file_size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
        print(f"✓ Size: {file_size_mb:.2f} MB")
        
        if file_size_mb > MAX_FILE_SIZE_MB:
            return jsonify({
                'error': f'File too large. Max size: {MAX_FILE_SIZE_MB}MB'
            }), 400
        
        # Import pdfplumber (faster than PyMuPDF for tables)
        try:
            import pdfplumber
        except ImportError:
            # Fallback to basic extraction
            return _extract_with_pymupdf(pdf_path, filename, tmpdir)
        
        # Open PDF with pdfplumber
        pdf = pdfplumber.open(pdf_path)
        total_pages = len(pdf.pages)
        pages_to_process = min(MAX_PAGES, total_pages)
        
        print(f"✓ Pages: {total_pages} (processing {pages_to_process})")
        
        # Create Excel file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Extracted Data'
        
        current_row = 1
        
        # Add header
        ws.cell(row=current_row, column=1, value=f"Extracted from: {filename}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Pages: {pages_to_process} of {total_pages}")
        ws.cell(row=current_row, column=1).font = Font(italic=True, size=10)
        current_row += 2
        
        # Process each page
        for page_num in range(pages_to_process):
            page = pdf.pages[page_num]
            
            # Page header
            ws.cell(row=current_row, column=1, value=f"Page {page_num + 1}")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E0E0E0", fill_type="solid")
            current_row += 1
            
            # Extract tables first (most important)
            tables = page.extract_tables()
            
            if tables:
                print(f"  Page {page_num + 1}: {len(tables)} tables found")
                
                for table_idx, table in enumerate(tables):
                    if not table:
                        continue
                    
                    # Table header
                    if len(tables) > 1:
                        ws.cell(row=current_row, column=1, value=f"Table {table_idx + 1}")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 1
                    
                    # Add table data
                    for row_idx, row in enumerate(table):
                        if not row:
                            continue
                        
                        for col_idx, cell in enumerate(row):
                            if cell:
                                ws.cell(row=current_row, column=col_idx + 1, value=str(cell).strip())
                                
                                # Header row styling
                                if row_idx == 0:
                                    ws.cell(row=current_row, column=col_idx + 1).font = Font(bold=True)
                                    ws.cell(row=current_row, column=col_idx + 1).fill = PatternFill(
                                        start_color="D9E1F2", fill_type="solid"
                                    )
                        
                        current_row += 1
                    
                    current_row += 1  # Spacing between tables
            
            else:
                # No tables found, extract text
                text = page.extract_text()
                
                if text:
                    print(f"  Page {page_num + 1}: Text extracted")
                    lines = text.strip().split('\n')
                    
                    for line in lines[:200]:  # Limit lines
                        if line.strip():
                            ws.cell(row=current_row, column=1, value=line.strip())
                            current_row += 1
                else:
                    ws.cell(row=current_row, column=1, value="[No content extracted]")
                    ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
                    current_row += 1
            
            current_row += 1  # Spacing between pages
            
            # Memory cleanup
            page = None
            gc.collect()
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Close PDF
        pdf.close()
        pdf = None
        gc.collect()
        
        # Save Excel
        excel_name = Path(filename).stem + '.xlsx'
        excel_path = os.path.join(tmpdir, excel_name)
        wb.save(excel_path)
        wb.close()
        
        print(f"✓ Excel created: {excel_name}")
        print(f"✓ Conversion complete in {pages_to_process} pages")
        print("="*60)
        
        # Send file
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=excel_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Conversion failed: {str(e)}'}), 500
        
    finally:
        # Cleanup
        if pdf:
            try:
                pdf.close()
            except:
                pass
        
        if tmpdir and os.path.exists(tmpdir):
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except:
                pass
        
        # Force memory cleanup
        gc.collect()
        gc.collect()


def _extract_with_pymupdf(pdf_path, filename, tmpdir):
    """Fallback extraction using PyMuPDF if pdfplumber not available"""
    print("Using PyMuPDF fallback...")
    
    import fitz
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    pages_to_process = min(MAX_PAGES, total_pages)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extracted Data'
    
    current_row = 1
    
    # Header
    ws.cell(row=current_row, column=1, value=f"Extracted from: {filename}")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 2
    
    # Extract text from each page
    for page_num in range(pages_to_process):
        page = doc[page_num]
        text = page.get_text()
        
        if text.strip():
            ws.cell(row=current_row, column=1, value=f"=== Page {page_num + 1} ===")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            lines = text.strip().split('\n')
            for line in lines[:200]:
                if line.strip():
                    ws.cell(row=current_row, column=1, value=line.strip())
                    current_row += 1
            
            current_row += 1
        
        gc.collect()
    
    doc.close()
    
    # Save
    excel_name = Path(filename).stem + '.xlsx'
    excel_path = os.path.join(tmpdir, excel_name)
    wb.save(excel_path)
    wb.close()
    
    return send_file(
        excel_path,
        as_attachment=True,
        download_name=excel_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
