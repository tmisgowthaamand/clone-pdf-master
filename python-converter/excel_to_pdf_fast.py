"""
Fast Excel/CSV to PDF converter optimized for cloud deployments (Render, Vercel)
Uses openpyxl + reportlab for pure Python conversion without external dependencies
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd


def convert_csv_to_excel_fast(csv_path):
    """Fast CSV to Excel conversion using pandas"""
    try:
        # Read CSV with pandas (handles various encodings)
        df = pd.read_csv(csv_path, encoding='utf-8')
        
        # Create Excel file
        excel_path = csv_path.rsplit('.', 1)[0] + '_converted.xlsx'
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        return excel_path
    except UnicodeDecodeError:
        # Try different encodings
        for encoding in ['latin-1', 'iso-8859-1', 'cp1252']:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                excel_path = csv_path.rsplit('.', 1)[0] + '_converted.xlsx'
                df.to_excel(excel_path, index=False, engine='openpyxl')
                return excel_path
            except:
                continue
    except Exception as e:
        print(f"CSV conversion error: {e}")
        return None


def excel_to_pdf_fast(excel_path, output_dir=None):
    """
    Fast Excel to PDF conversion using pure Python (openpyxl + reportlab)
    Optimized for cloud deployments - no external dependencies
    """
    if output_dir is None:
        output_dir = os.path.dirname(excel_path) or '.'
    
    print(f"\n{'='*60}")
    print(f"FAST EXCEL TO PDF CONVERSION (Cloud-Optimized)")
    print(f"Input: {excel_path}")
    print(f"{'='*60}\n")
    
    # Load workbook
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active
    
    # Get data dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"Sheet: {ws.title}")
    print(f"Dimensions: {max_row} rows × {max_col} columns")
    
    # Extract data into list of lists
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_data = []
        for cell in row:
            value = cell.value
            if value is None:
                row_data.append('')
            else:
                row_data.append(str(value))
        data.append(row_data)
    
    wb.close()
    
    # Determine page orientation based on columns
    if max_col > 6:
        pagesize = landscape(A4)
        print("Using landscape orientation (wide table)")
    else:
        pagesize = A4
        print("Using portrait orientation")
    
    # Create PDF
    pdf_name = Path(excel_path).stem + '.pdf'
    pdf_path = Path(output_dir) / pdf_name
    
    c = canvas.Canvas(str(pdf_path), pagesize=pagesize)
    width, height = pagesize
    
    # Calculate column widths dynamically
    available_width = width - 40  # 20pt margins on each side
    col_widths = [available_width / max_col] * max_col
    
    # Adjust column widths based on content
    for col_idx in range(max_col):
        max_len = 0
        for row_idx in range(min(max_row, 100)):  # Check first 100 rows
            if row_idx < len(data) and col_idx < len(data[row_idx]):
                max_len = max(max_len, len(data[row_idx][col_idx]))
        
        # Estimate width (rough approximation)
        estimated_width = min(max_len * 6, available_width / 2)
        col_widths[col_idx] = max(estimated_width, available_width / (max_col * 2))
    
    # Normalize widths to fit page
    total_width = sum(col_widths)
    if total_width > available_width:
        scale = available_width / total_width
        col_widths = [w * scale for w in col_widths]
    
    # Create table with ReportLab
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Style the table
    style = TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ])
    
    table.setStyle(style)
    
    # Calculate table dimensions and draw
    table_width, table_height = table.wrap(available_width, height - 60)
    
    # Position table
    x = 20
    y = height - 40 - table_height
    
    if y < 20:  # Table too tall for one page
        # Split into multiple pages
        rows_per_page = int((height - 80) / (table_height / max_row))
        
        for page_start in range(0, max_row, rows_per_page):
            page_end = min(page_start + rows_per_page, max_row)
            
            # Include header row on each page
            if page_start > 0:
                page_data = [data[0]] + data[page_start:page_end]
            else:
                page_data = data[page_start:page_end]
            
            page_table = Table(page_data, colWidths=col_widths, repeatRows=1)
            page_table.setStyle(style)
            
            page_table_width, page_table_height = page_table.wrap(available_width, height - 60)
            page_table.drawOn(c, x, height - 40 - page_table_height)
            
            if page_end < max_row:
                c.showPage()
    else:
        # Single page
        table.drawOn(c, x, y)
    
    c.save()
    
    print(f"\n[OK] PDF created: {pdf_path}")
    print(f"Size: {os.path.getsize(pdf_path) / 1024:.1f} KB")
    print(f"{'='*60}\n")
    
    return str(pdf_path)


def convert_excel_to_pdf_optimized(excel_path, output_dir=None):
    """
    Main conversion function with CSV support
    Optimized for cloud deployments
    """
    # Handle CSV files
    if excel_path.lower().endswith('.csv'):
        print("Converting CSV to Excel first...")
        excel_path = convert_csv_to_excel_fast(excel_path)
        if not excel_path:
            raise ValueError("CSV conversion failed")
    
    # Convert to PDF
    return excel_to_pdf_fast(excel_path, output_dir)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_dir = sys.argv[2] if len(sys.argv) > 2 else None
        result = convert_excel_to_pdf_optimized(input_file, output_dir)
        print(f"Result: {result}")
    else:
        print("Usage: python excel_to_pdf_fast.py <input_file> [output_dir]")
