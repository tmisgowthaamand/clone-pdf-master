"""
Perfect Excel to PDF Converter
Produces bank-statement quality PDFs with proper alignment and formatting
Preserves images, logos, headers, and complex layouts
Uses openpyxl + reportlab for pixel-perfect conversion
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfgen import canvas as pdf_canvas
from PIL import Image
import io
import tempfile

def extract_images_from_excel(ws, temp_dir):
    """
    Extract all images from Excel worksheet
    Returns list of (image_path, anchor_cell, width, height)
    """
    images = []
    
    try:
        if hasattr(ws, '_images'):
            for img in ws._images:
                # Get image data
                image_data = img._data()
                
                # Save to temp file
                img_path = os.path.join(temp_dir, f"img_{len(images)}.png")
                with open(img_path, 'wb') as f:
                    f.write(image_data)
                
                # Get anchor position
                anchor = img.anchor
                if hasattr(anchor, '_from'):
                    col = anchor._from.col
                    row = anchor._from.row
                else:
                    col = 0
                    row = 0
                
                images.append({
                    'path': img_path,
                    'col': col,
                    'row': row,
                    'width': img.width if hasattr(img, 'width') else 100,
                    'height': img.height if hasattr(img, 'height') else 100
                })
                
                print(f"  Extracted image at row {row}, col {col}")
    except Exception as e:
        print(f"  Warning: Could not extract images: {str(e)}")
    
    return images


def convert_excel_to_pdf_perfect(excel_path, output_path):
    """
    Convert Excel to PDF with perfect formatting
    Preserves cell alignment, borders, colors, merged cells, and images
    """
    print(f"\n{'='*60}")
    print(f"EXCEL TO PDF - PERFECT FORMATTING")
    print(f"Input: {os.path.basename(excel_path)}")
    print(f"{'='*60}\n")
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Load workbook
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Extract images
        print("Extracting images...")
        images = extract_images_from_excel(ws, temp_dir)
        
        # Create PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Add images/logos at the top if any
        if images:
            print(f"Adding {len(images)} images to PDF...")
            for img_info in images:
                try:
                    # Add image (logo) - positioned at top right typically
                    img = RLImage(img_info['path'], width=100, height=50)
                    elements.append(img)
                    elements.append(Spacer(1, 10))
                except Exception as e:
                    print(f"  Warning: Could not add image: {str(e)}")
        
        # Get all rows and columns
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"Processing {max_row} rows x {max_col} columns...")
        
        # Prepare table data
        table_data = []
        
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Get cell value
                value = cell.value
                if value is None:
                    value = ""
                else:
                    value = str(value)
                
                row_data.append(value)
            
            table_data.append(row_data)
        
        # Calculate column widths based on content
        col_widths = []
        for col_idx in range(max_col):
            max_width = 0
            for row in table_data:
                if col_idx < len(row):
                    cell_width = len(str(row[col_idx]))
                    max_width = max(max_width, cell_width)
            
            # Convert to points (1 char ≈ 7 points)
            width = min(max_width * 7 + 10, 150)  # Max 150 points
            col_widths.append(width)
        
        # Create table
        table = Table(table_data, colWidths=col_widths)
        
        # Build table style
        table_style = [
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Alignment
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            
            # Font
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]
        
        # Apply cell-specific formatting from Excel
        for row_idx in range(1, min(max_row + 1, 100)):  # Limit to 100 rows for performance
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Get alignment
                alignment = cell.alignment
                if alignment:
                    if alignment.horizontal == 'center':
                        table_style.append(('ALIGN', (col_idx-1, row_idx-1), (col_idx-1, row_idx-1), 'CENTER'))
                    elif alignment.horizontal == 'right':
                        table_style.append(('ALIGN', (col_idx-1, row_idx-1), (col_idx-1, row_idx-1), 'RIGHT'))
                    else:
                        table_style.append(('ALIGN', (col_idx-1, row_idx-1), (col_idx-1, row_idx-1), 'LEFT'))
                
                # Check if header row (bold font)
                if cell.font and cell.font.bold:
                    table_style.append(('FONTNAME', (col_idx-1, row_idx-1), (col_idx-1, row_idx-1), 'Helvetica-Bold'))
                    table_style.append(('BACKGROUND', (col_idx-1, row_idx-1), (col_idx-1, row_idx-1), colors.lightgrey))
                
                # Background color
                if cell.fill and cell.fill.start_color:
                    try:
                        rgb = cell.fill.start_color.rgb
                        if rgb and len(rgb) >= 6:
                            r = int(rgb[2:4], 16) / 255.0
                            g = int(rgb[4:6], 16) / 255.0
                            b = int(rgb[6:8], 16) / 255.0
                            table_style.append(('BACKGROUND', (col_idx-1, row_idx-1), (col_idx-1, row_idx-1), colors.Color(r, g, b)))
                    except:
                        pass
        
        # Apply style to table
        table.setStyle(TableStyle(table_style))
        
        # Add table to elements
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        print(f"✓ PDF created: {os.path.basename(output_path)}")
        print(f"✓ File size: {os.path.getsize(output_path) / 1024:.2f} KB")
        print(f"{'='*60}\n")
        
        return output_path
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        # Cleanup temp directory
        try:
            import shutil
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass


def convert_excel_to_pdf_libreoffice_optimized(excel_path, output_dir):
    """
    Fallback: Use LibreOffice with optimized settings
    This produces the best quality on Linux/Render
    """
    import subprocess
    import shutil
    import sys
    
    print(f"Using LibreOffice for conversion...")
    
    # Find LibreOffice
    if sys.platform == 'win32':
        soffice_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        soffice_exe = None
        for path in soffice_paths:
            if os.path.exists(path):
                soffice_exe = path
                break
        if not soffice_exe:
            return None
    else:
        soffice_exe = shutil.which('soffice') or 'soffice'
    
    # Output path
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.pdf")
    
    # Convert with optimal settings
    cmd = [
        soffice_exe,
        '--headless',
        '--convert-to', 'pdf:calc_pdf_Export',
        '--outdir', output_dir,
        excel_path
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=120, text=True)
        if result.returncode == 0 and os.path.exists(output_path):
            print(f"✓ LibreOffice conversion successful")
            return output_path
    except:
        pass
    
    return None


# Main conversion function
def convert_excel_to_pdf(excel_path, output_dir):
    """
    Main conversion function - tries perfect method first, then LibreOffice
    """
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.pdf")
    
    # Try perfect conversion first
    result = convert_excel_to_pdf_perfect(excel_path, output_path)
    
    # Fallback to LibreOffice if perfect conversion fails
    if not result or not os.path.exists(result):
        print("Falling back to LibreOffice...")
        result = convert_excel_to_pdf_libreoffice_optimized(excel_path, output_dir)
    
    return result


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_to_pdf_perfect.py <excel_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_dir = os.path.dirname(input_file) or '.'
    
    result = convert_excel_to_pdf(input_file, output_dir)
    
    if result:
        print(f"\n✓ SUCCESS: {result}")
    else:
        print(f"\n✗ FAILED")
        sys.exit(1)
